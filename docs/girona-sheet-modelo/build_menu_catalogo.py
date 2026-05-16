#!/usr/bin/env python3
"""
Genera menu_catalogo_girona.csv: menú ↔ recetas ↔ costo por ingredientes (JSON).

Columnas:
  ID Menú | Nombre Producto Venta | Categoría Menú | Área | ID Receta | Precio Venta
  | Costo Producto | Tiene ingredientes costeo | ID Backend

Área Bar = mismas categorías que la pestaña Bar en Menu/index.tsx, más «Postres» cuando
aplica (mismo criterio que menu_escrito en seed_girona_data).

ID Receta:
  REC-001… = orden en recetario_bar.json
  REC-037+ = platos cocina (REST_DISHES_ORDERED) que no choquen con el bar
  siguientes = ítems bar extra (p. ej. cervezas desde menu_escrito) y luego cualquier
  nombre que venga solo por API/export.

Costo producto: suma (weight × price) cuando ingredients trae objetos con peso y precio;
si solo hay texto en la lista, Costo queda vacío pero «Tiene ingredientes costeo» = SI.

El menú «actual» (61 Restaurante + 57 Bar) no está guardado en Git: es el de tu BD o del
Excel «Menú Girona Picnic» (hojas Restaurante y Bar, ver sync_menu_picnic_xlsx.py).

  1) Volcado desde el mismo endpoint que el front:
     curl -sS "http://HOST:8000/menu/items" > docs/girona-sheet-modelo/menu_items_export.json
     python3 docs/girona-sheet-modelo/build_menu_catalogo.py

  2) Sin backend: copiá el .xlsx a esta carpeta o usá --picnic-xlsx / PICNIC_MENU_XLSX.

Orden de fuentes: menu_items_export.json → GET /menu/items → Excel Picnic → fallback repo.

Si existe GIRONA_DATA_DIR con menu_escrito.txt, el último fallback puede sumar ítems del bar escrito.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.request
from decimal import Decimal
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
OUT = Path(__file__).resolve().parent / "menu_catalogo_girona.csv"
RECETARIO_JSON = ROOT / "girona-back/app/recetario_bar.json"
EXPORT_JSON = Path(__file__).resolve().parent / "menu_items_export.json"
SCRIPTS_DIR = ROOT / "girona-front/scripts"
# Si colocás aquí el Excel de carta (mismo formato que sync_menu_picnic_xlsx), se usa sin variables de entorno.
PICNIC_XLSX_CANDIDATES = [
    Path(__file__).resolve().parent / "Menu_Girona_Picnic.xlsx",
    Path(__file__).resolve().parent / "Menu_Girona_Picnic (1).xlsx",
]

BAR_CATEGORY_KEYS = frozenset(
    {
        "bebidas",
        "malteadas",
        "dulces bar",
        "sodas",
        "gaseosas",
        "para el almuerzo",
        "cervezas nacionales",
        "cervezas internacionales",
        "micheladas",
        "licores y shots",
        "cubetazos",
        "cocteleria",
        "vinos",
        "postres",
    },
)


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s.strip()) if unicodedata.category(c) != "Mn"
    )


def category_key(value: str) -> str:
    return strip_accents(value).lower().strip()


def is_bar_area(category: str) -> bool:
    return category_key(category) in BAR_CATEGORY_KEYS


def norm_name(value: str) -> str:
    return re.sub(r"\s+", " ", strip_accents(value).lower().strip())


def display_category(cat: str) -> str:
    t = cat.strip()
    if not t:
        return t
    return t[0].upper() + t[1:]


def find_menu_escrito() -> Path | None:
    env = os.environ.get("GIRONA_DATA_DIR")
    if env:
        p = Path(env).resolve() / "menu_escrito.txt"
        if p.is_file():
            return p
    return None


def resolve_picnic_xlsx(cli_path: str | None) -> Path | None:
    if cli_path:
        p = Path(cli_path).expanduser().resolve()
        return p if p.is_file() else None
    env = os.environ.get("PICNIC_MENU_XLSX")
    if env:
        p = Path(env).expanduser().resolve()
        return p if p.is_file() else None
    for p in PICNIC_XLSX_CANDIDATES:
        if p.is_file():
            return p.resolve()
    return None


def _normalize_category_label_picnic(raw: str) -> str:
    c = (raw or "").strip()
    return re.sub(r"\s*&\s*", " y ", c)


def _as_decimal_picnic(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(" ", "").replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _parse_picnic_sheet_rows(ws) -> list[tuple[str, str, str | None, Decimal, str | None]]:
    out: list[tuple[str, str, str | None, Decimal, str | None]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row) + [None] * (5 - len(row))
        cat, name, desc, price_raw, avail = cells[0], cells[1], cells[2], cells[3], cells[4]
        if name is None or not str(name).strip():
            continue
        name_s = " ".join(str(name).split())
        cat_s = str(cat).strip() if cat is not None else ""
        if not cat_s:
            continue
        out.append(
            (
                cat_s,
                name_s,
                None if desc is None or str(desc).strip() == "" else str(desc).strip(),
                _as_decimal_picnic(price_raw),
                None if avail is None or str(avail).strip() == "" else str(avail).strip(),
            )
        )
    return out


def load_menu_from_picnic_xlsx(path: Path) -> tuple[list[dict], str]:
    """
    Misma forma que sync_menu_picnic_xlsx: hojas «Bar» y «Restaurante», columnas A–E.
    Área = hoja (no se infiere solo por categoría).
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "Instalá openpyxl para leer el Excel (p. ej. pip install openpyxl "
            "o un entorno con girona-back/requirements.txt)."
        ) from e
    if not path.is_file():
        raise FileNotFoundError(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items: list[dict] = []
    pos = 0
    for sheet_name, is_restaurant in (("Bar", False), ("Restaurante", True)):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"La hoja «{sheet_name}» no existe en {path.name}")
        for cat_raw, name_s, _desc, price_dec, _avail in _parse_picnic_sheet_rows(wb[sheet_name]):
            cat_fmt = display_category(_normalize_category_label_picnic(cat_raw))
            q = price_dec
            price_val: int | float = int(q) if q == q.to_integral_value() else float(q)
            items.append(
                {
                    "id": None,
                    "name": name_s,
                    "category": cat_fmt,
                    "price": price_val,
                    "ingredients": None,
                    "_area_sheet": "Restaurante" if is_restaurant else "Bar",
                    "_src_order": pos,
                }
            )
            pos += 1
    return items, f"Excel Picnic {path.name}"


def bar_categorize_seed(name: str) -> tuple[str, str]:
    """Copia de la lógica de _bar_categorize en seed_girona_data (bar en menu_escrito)."""
    t = " ".join(name.split())
    n = t.lower()
    if any(
        k in n
        for k in (
            "cerveza",
            "poker",
            "heineken",
            "stella",
            "aguila",
            "coron",
            "club col",
            "postob",
            "ginet",
            "andina",
            "segu ole",
        )
    ):
        return ("Cervezas nacionales", t)
    if any(
        k in n
        for k in (
            "ginebra",
            "tequila",
            "ron",
            "vino",
            "aguard",
            "champag",
            "triple",
            "soda tón",
            "coca",
            "gaseos",
            "blue",
            "tónica",
            "cuatro",
        )
    ):
        if "vino" in n:
            return ("Vinos", t)
        if any(
            k in n
            for k in (
                "coca",
                "gaseos",
                "ton",
                "soda",
                "hola",
                "andina",
                "segu",
                "ginet",
                "postob",
                "manzana",
                "uva",
                "cola y pola",
            )
        ):
            return ("Sodas", t)
        return ("Cocteleria", t)
    if any(k in n for k in ("helado", "gomit", "banderit", "chispa", "leche conden", "tajín", "chantilly")):
        return ("Postres", t)
    return ("Sodas", t) if t else ("Sodas", t)


def parse_menu_escrito_bar(path: Path) -> list[tuple[str, str]]:
    raw = path.read_text(encoding="utf-8", errors="replace").splitlines()
    bar: list[tuple[str, str]] = []
    mode: str = "body"
    for i, line in enumerate(raw, start=1):
        t = " ".join(line.split()).strip()
        if not t or t == "MENÚ":
            continue
        if "Inventario cristalería" in t:
            mode = "cristal" if "21 de enero" in t else "cristal_drop"
            continue
        if mode == "cristal_drop":
            continue
        if mode == "cristal":
            continue
        if 94 <= i <= 132:
            bar.append(bar_categorize_seed(t))
    return bar


def balanced_range(s: str, start_idx: int, open_c: str, close_c: str) -> tuple[int, int] | None:
    if start_idx >= len(s) or s[start_idx] != open_c:
        return None
    depth = 0
    i = start_idx
    while i < len(s):
        c = s[i]
        if c == open_c:
            depth += 1
        elif c == close_c:
            depth -= 1
            if depth == 0:
                return start_idx, i
        i += 1
    return None


def split_top_level_js_objects(arr_inner: str) -> list[str]:
    objs: list[str] = []
    depth = 0
    start: int | None = None
    for i, c in enumerate(arr_inner):
        if c == "{":
            if depth == 0:
                start = i
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0 and start is not None:
                objs.append(arr_inner[start : i + 1])
                start = None
    return objs


def parse_ingredient_object_blob(blob: str) -> dict | None:
    m = re.search(r'name:\s*"([^"]*)"', blob)
    if not m:
        return None
    name = m.group(1)
    wm = re.search(r"weight:\s*([\d.]+)", blob)
    pm = re.search(r"price:\s*([\d.]+)", blob)
    if not wm or not pm:
        return None
    try:
        return {"name": name, "weight": float(wm.group(1)), "price": float(pm.group(1))}
    except ValueError:
        return None


def load_mjs_menu_recipes() -> dict[str, list[dict]]:
    """menuName + ingredients desde upsert-*-recipes.mjs (objetos con name, weight, price)."""
    out: dict[str, list[dict]] = {}
    if not SCRIPTS_DIR.is_dir():
        return out
    for path in sorted(SCRIPTS_DIR.glob("upsert-*-recipes.mjs")):
        text = path.read_text(encoding="utf-8")
        pos = 0
        while True:
            m = re.search(r'menuName:\s*"([^"]+)"', text[pos:])
            if not m:
                break
            abs_pos = pos + m.end()
            idx = text.find("ingredients:", abs_pos)
            if idx < 0 or idx - abs_pos > 1200:
                pos = abs_pos
                continue
            lb = text.find("[", idx)
            br = balanced_range(text, lb, "[", "]")
            if not br:
                pos = abs_pos
                continue
            inner = text[br[0] + 1 : br[1]]
            ingredients: list[dict] = []
            for blob in split_top_level_js_objects(inner):
                row = parse_ingredient_object_blob(blob)
                if row:
                    ingredients.append(row)
            if ingredients:
                out[norm_name(m.group(1))] = ingredients
            pos = br[1] + 1
    return out


def load_recetario_bar_names() -> list[str]:
    if not RECETARIO_JSON.is_file():
        return []
    raw = json.loads(RECETARIO_JSON.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        return []
    out: list[str] = []
    for x in raw:
        if isinstance(x, dict):
            n = (x.get("name") or "").strip()
            if n:
                out.append(n)
    return out


REST_DISHES_ORDERED: list[str] = [
    "PECHUGA EN SALSA DE MARACUYA",
    "MOJARRA",
    "CHICHARRON AL BARRIL",
    "CHICHARRON EN REDUCCION DE PANELA",
    "SALMON AL PESTO",
    "PICADA PARA CUATRO",
    "CLASICA BURGUER",
    "GIRONA BURGUER",
    "CRSIPY BURGUER",
    "CAFFETO BURGUER",
    "KETO BURGUER",
    "PHILLY POWER BURGER",
    "LA MÚNICH",
    "LA MEDALLO",
    "LA CHULA",
    "LA QUESUDA BURGUER",
    "DESGRANADO GIRONA",
    "PAPAS GIRONA",
    "QUESADILLA DE CARNE",
    "QUESADILLA DE POLLO",
    "QUESADILLA MIXTA",
    "SALCHIPAPA JR",
    "PERRO GIRONÉS",
    "PERRO BÁRBARO",
    "PERRO CHILI-DOG",
    "PERRO FORASTERO",
    "TORRE GIRONA",
    "FETUCCINI CARBONARA",
    "GREGORIANA MAR Y TIERRA",
    "ENSALADA GIRONA",
    "CEVICHE DE CHICHARRÓN",
    "SALMÓN EN REDUCCIÓN DE FRUTOS ROJOS",
    "PULPITO DE SALCHICHA",
    "NUGGETS DE POLLO",
    "MIGAO GIRONA",
    "CUAJADA CON REDUCCIÓN EN PANELA",
    "CREPES DULCE TENTACIÓN",
    "WAFFLES",
]


def costo_desde_ingredientes(ingredients: object) -> str:
    if not isinstance(ingredients, list) or not ingredients:
        return ""
    total = Decimal("0")
    for row in ingredients:
        if not isinstance(row, dict):
            continue
        try:
            w = Decimal(str(row.get("weight") or 0))
            p = Decimal(str(row.get("price") or 0))
        except Exception:
            continue
        total += w * p
    if total == 0:
        return ""
    return str(int(total.quantize(Decimal("1"))))


def tiene_ingredientes_asociados(ingredients: object) -> str:
    """SI si hay lista de ingredientes (texto o filas costeables); alinea con menú ↔ receta."""
    if not isinstance(ingredients, list) or not ingredients:
        return "NO"
    return "SI"


def fetch_menu_items(api_base: str) -> list[dict] | None:
    base = api_base.rstrip("/")
    url = f"{base}/menu/items"
    try:
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw_txt = resp.read().decode("utf-8")
        if not raw_txt.strip() or raw_txt.lstrip().startswith("<"):
            print(
                f"Aviso: {url} no devolvió JSON (respuesta vacía u HTML ¿otro servicio en el puerto?).",
                file=sys.stderr,
            )
            return None
        data = json.loads(raw_txt)
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, OSError) as e:
        print(f"Aviso: no se pudo GET {url}: {e}", file=sys.stderr)
        return None
    if not isinstance(data, list):
        return None
    return [x for x in data if isinstance(x, dict)]


def static_fallback_menu() -> list[dict]:
    by_norm: dict[str, dict] = {}
    mjs = load_mjs_menu_recipes()

    escrito = find_menu_escrito()
    if escrito:
        for cat, name in parse_menu_escrito_bar(escrito):
            k = norm_name(name)
            by_norm[k] = {
                "id": None,
                "name": name,
                "category": cat,
                "price": 8000,
                "ingredients": None,
            }

    if RECETARIO_JSON.is_file():
        for entry in json.loads(RECETARIO_JSON.read_text(encoding="utf-8")):
            if not isinstance(entry, dict):
                continue
            name = (entry.get("name") or "").strip()
            if not name:
                continue
            cat = (entry.get("category") or "Bebidas").strip() or "Bebidas"
            price = entry.get("price")
            try:
                p = int(Decimal(str(price))) if price not in (None, "") else 0
            except Exception:
                p = 0
            k = norm_name(name)
            by_norm[k] = {
                "id": None,
                "name": name,
                "category": cat,
                "price": p,
                "ingredients": entry.get("ingredients"),
            }

    for dish in REST_DISHES_ORDERED:
        k = norm_name(dish)
        if k in by_norm:
            continue
        by_norm[k] = {
            "id": None,
            "name": dish,
            "category": "Menú restaurante",
            "price": 0,
            "ingredients": mjs.get(k),
        }
    if not by_norm:
        return []
    return list(by_norm.values())


def load_export_json() -> list[dict] | None:
    if not EXPORT_JSON.is_file():
        return None
    try:
        data = json.loads(EXPORT_JSON.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        print(f"Aviso: JSON inválido en {EXPORT_JSON}: {e}", file=sys.stderr)
        return None
    if not isinstance(data, list):
        return None
    return [x for x in data if isinstance(x, dict)]


def parse_price(raw: object) -> int:
    if raw is None:
        return 0
    if isinstance(raw, bool):
        return 0
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        return int(round(raw))
    s = str(raw).strip()
    if not s:
        return 0
    cleaned = re.sub(r"[^\d.,-]", "", s.replace(" ", ""))
    if not cleaned:
        return 0
    try:
        return int(Decimal(cleaned.replace(",", ".")))
    except Exception:
        return 0


def build_receta_map(ordered_rows: list[dict]) -> dict[str, str]:
    """
    Asigna REC-### estable: recetario → platos REST_DISHES → bar extra → resto en orden de salida.
    """
    mp: dict[str, str] = {}
    n = 1
    for name in load_recetario_bar_names():
        mp[norm_name(name)] = f"REC-{n:03d}"
        n += 1
    next_id = 37
    for nm in sorted(REST_DISHES_ORDERED, key=str.lower):
        k = norm_name(nm)
        if k not in mp:
            mp[k] = f"REC-{next_id:03d}"
            next_id += 1

    rec_bar_keys = {norm_name(x) for x in load_recetario_bar_names()}
    bar_extras: list[str] = []
    seen_x: set[str] = set()
    for row in ordered_rows:
        if row["area"] != "Bar":
            continue
        k = norm_name(row["name"])
        if k in rec_bar_keys or k in seen_x:
            continue
        seen_x.add(k)
        bar_extras.append(row["name"])
    for nm in sorted(bar_extras, key=str.lower):
        k = norm_name(nm)
        if k not in mp:
            mp[k] = f"REC-{next_id:03d}"
            next_id += 1

    for row in ordered_rows:
        k = norm_name(row["name"])
        if k not in mp:
            mp[k] = f"REC-{next_id:03d}"
            next_id += 1
    return mp


def enrich_with_mjs(items: list[dict], mjs: dict[str, list[dict]]) -> None:
    for it in items:
        ing = it.get("ingredients")
        if ing:
            continue
        k = norm_name(str(it.get("name") or ""))
        if k in mjs:
            it["ingredients"] = mjs[k]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--api-url", default=None, help="Base URL backend (sin /api; default env BACKEND_URL)")
    ap.add_argument(
        "--picnic-xlsx",
        default=None,
        help="Excel Menú Girona Picnic (hojas Bar + Restaurante). Si no se indica, se usa "
        "PICNIC_MENU_XLSX o un .xlsx en la misma carpeta que este script.",
    )
    args = ap.parse_args()
    api_base = (args.api_url or os.environ.get("BACKEND_URL") or "http://127.0.0.1:8000").strip()

    mjs = load_mjs_menu_recipes()

    items: list[dict] | None = None
    source = ""
    picnic_path = resolve_picnic_xlsx(args.picnic_xlsx)

    items = load_export_json()
    if items is not None:
        source = f"archivo {EXPORT_JSON.name}"
    if items is None:
        items = fetch_menu_items(api_base)
        if items is not None:
            source = f"API {api_base}/menu/items"
    if items is None and picnic_path is not None:
        try:
            items, source = load_menu_from_picnic_xlsx(picnic_path)
        except (OSError, ValueError, ImportError) as e:
            print(f"Aviso: no se pudo leer Excel Picnic {picnic_path}: {e}", file=sys.stderr)
            items = None
    if not items:
        items = static_fallback_menu()
        bits = ["recetario_bar.json", "scripts upsert-*-recipes.mjs"]
        if find_menu_escrito():
            bits.insert(0, "menu_escrito.txt (GIRONA_DATA_DIR)")
        source = "fallback repo (" + ", ".join(bits) + ")"
    else:
        enrich_with_mjs(items, mjs)

    preserve_order = any(isinstance(it, dict) and "_src_order" in it for it in items)

    normalized: list[dict] = []
    for it in items:
        if not isinstance(it, dict):
            continue
        name = str(it.get("name") or "").strip()
        if not name:
            continue
        cat = str(it.get("category") or "").strip() or "Sin categoría"
        precio = parse_price(it.get("price"))
        ing = it.get("ingredients")
        bid = it.get("id")
        area_override = it.get("_area_sheet")
        if isinstance(area_override, str) and area_override in ("Bar", "Restaurante"):
            area = area_override
        else:
            area = "Bar" if is_bar_area(cat) else "Restaurante"
        row_d: dict = {
            "id": bid,
            "name": name,
            "category": cat,
            "price": precio,
            "ingredients": ing,
            "area": area,
        }
        if preserve_order and "_src_order" in it:
            row_d["_src_order"] = int(it["_src_order"])
        normalized.append(row_d)

    if preserve_order:
        ordered = sorted(normalized, key=lambda x: x.get("_src_order", 0))
        for r in ordered:
            r.pop("_src_order", None)
    else:
        bar = [x for x in normalized if x["area"] == "Bar"]
        rest = [x for x in normalized if x["area"] != "Bar"]
        bar.sort(key=lambda x: (category_key(x["category"]), x["name"].lower()))
        rest.sort(key=lambda x: (category_key(x["category"]), x["name"].lower()))
        ordered = bar + rest

    rec_map = build_receta_map(ordered)

    bar = [x for x in ordered if x["area"] == "Bar"]
    rest = [x for x in ordered if x["area"] != "Bar"]

    headers = [
        "ID Menú",
        "Nombre Producto Venta",
        "Categoría Menú",
        "Área",
        "ID Receta",
        "Precio Venta",
        "Costo Producto",
        "Tiene ingredientes costeo",
        "ID Backend",
    ]

    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for i, row in enumerate(ordered, start=1):
            rid = rec_map.get(norm_name(row["name"]), "")
            costo = costo_desde_ingredientes(row["ingredients"])
            w.writerow(
                [
                    f"MEN-{i:03d}",
                    row["name"],
                    display_category(row["category"]),
                    row["area"],
                    rid,
                    row["price"],
                    costo,
                    tiene_ingredientes_asociados(row["ingredients"]),
                    row["id"] if row["id"] is not None else "",
                ]
            )

    print(f"Fuente: {source}")
    print(f"Escrito {OUT}: total={len(ordered)} (Bar={len(bar)}, Restaurante={len(rest)}).")
    if len(bar) != 57 or len(rest) != 61:
        print(
            "Nota esperada: 57 ítems en pestaña Bar y 61 en Restaurante (según carta Picnic/BD).",
            file=sys.stderr,
        )
        print(
            f"  Obtén el menú real con: curl -sS '{api_base.rstrip('/')}/menu/items' > {EXPORT_JSON} "
            f"o copiá el Excel Picnic en docs/girona-sheet-modelo/ y definí PICNIC_MENU_XLSX.",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
