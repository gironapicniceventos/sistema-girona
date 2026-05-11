"""
Actualiza el menú (tabla `menu_items`) desde el Excel «Menú Girona Picnic»
(hojas Restaurante y Bar).

- Ajusta nombre, categoría, precio y descripción según el archivo.
- No modifica `inventory_products`, recetas ni movimientos.
- Conserva el JSON `ingredients` de cada ítem existente.
- Con `--replace`, desactiva (`is_active=false`) todo ítem activo que no esté en el Excel,
  excepto recetas internas (categoría Insumos, Recetas, nombres `[Insumo]…`). No borra filas
  (historial POS/ventas sigue enlazado). Inventario no se modifica.

Uso (desde girona-back):

  python -m app.sync_menu_picnic_xlsx "/ruta/a/Menu_Girona_Picnic (1).xlsx"
  python -m app.sync_menu_picnic_xlsx --replace "/ruta/a/Menu_Girona_Picnic (1).xlsx"

Carga `DATABASE_URL` desde el entorno. Si no está definida, intenta leer
`.env` en la raíz del repo (junto a docker-compose) y `girona-back/.env`,
sin sobrescribir variables que ya hubieras exportado.

O:

  export PICNIC_MENU_XLSX="/ruta/al/archivo.xlsx"
  python -m app.sync_menu_picnic_xlsx
"""
from __future__ import annotations

import os
import re
import sys
import unicodedata
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl


def _load_env_from_dotenv_files() -> None:
    """Lee .env sin dependencias extra; no pisa variables ya definidas en el shell."""
    here = Path(__file__).resolve()
    for path in (here.parents[2] / ".env", here.parents[1] / ".env"):
        if not path.is_file():
            continue
        for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("export "):
                line = line[7:].strip()
            if "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            if not key or key in os.environ:
                continue
            value = value.strip()
            if len(value) >= 2 and value[0] == value[-1] and value[0] in "\"'":
                value = value[1:-1]
            os.environ[key] = value


def _ensure_database_url_from_compose_env() -> None:
    """
    En la raíz del repo, docker-compose suele definir POSTGRES_* pero no DATABASE_URL
    para el host; el backend sólo recibe DATABASE_URL dentro del contenedor.
    """
    if os.getenv("DATABASE_URL"):
        return
    user = os.getenv("POSTGRES_USER")
    password = os.getenv("POSTGRES_PASSWORD", "")
    db = os.getenv("POSTGRES_DB")
    port = os.getenv("POSTGRES_HOST_PORT", "25432")
    if user and db:
        from urllib.parse import quote_plus

        if password != "":
            auth = f"{quote_plus(user)}:{quote_plus(password)}"
        else:
            auth = quote_plus(user)
        os.environ["DATABASE_URL"] = (
            f"postgresql://{auth}@127.0.0.1:{port}/{db}"
        )


_load_env_from_dotenv_files()
_ensure_database_url_from_compose_env()

from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session

from . import db, models

# Misma lógica que la pestaña Bar del POS (pos-screen.tsx), más variantes del Excel.
_BAR_CATEGORY_KEYS = frozenset(
    {
        "bebidas",
        "sodas",
        "gaseosas",
        "para el almuerzo",
        "cervezas nacionales",
        "cervezas internacionales",
        "micheladas",
        "licores y shots",
        "licores & shots",
        "cubetazos",
        "cocteleria",
        "vinos",
    }
)


def _strip_accents(value: str) -> str:
    s = unicodedata.normalize("NFD", (value or "").strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _pos_category_key(value: str) -> str:
    return _strip_accents(value).lower().strip()


def _norm_name(value: str) -> str:
    return " ".join(_strip_accents(value).lower().split())


def _normalize_category_label(raw: str) -> str:
    c = (raw or "").strip()
    c = re.sub(r"\s*&\s*", " y ", c)
    return c


def _format_category(value: str) -> str:
    value = value.strip()
    if not value:
        return value
    return value[:1].upper() + value[1:]


def _is_bar_category(category: str) -> bool:
    return _pos_category_key(category) in _BAR_CATEGORY_KEYS


def _is_internal_menu_item(m: models.MenuItem) -> bool:
    """Ítems de costeo / recetas internas: no se quitan con --replace."""
    cat = _norm_name(m.category)
    if cat in ("insumos", "recetas"):
        return True
    if (m.name or "").strip().upper().startswith("[INSUMO]"):
        return True
    return False


def _as_decimal(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(" ", "").replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _build_description(desc: Any, availability: Any) -> str | None:
    parts: list[str] = []
    if desc is not None and str(desc).strip():
        parts.append(str(desc).strip())
    if availability is not None and str(availability).strip():
        parts.append(f"Disponibilidad: {str(availability).strip()}")
    if not parts:
        return None
    return "\n\n".join(parts)


def _parse_sheet_rows(ws) -> list[tuple[str, str, str | None, Decimal, str | None]]:
    out: list[tuple[str, str, str | None, Decimal, str | None]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row) + [None] * (5 - len(row))
        cat, name, desc, price_raw, avail = cells[0], cells[1], cells[2], cells[3], cells[4]
        if name is None or not str(name).strip():
            continue
        name_s = " ".join(str(name).split())
        cat_s = str(cat).strip() if cat is not None else ""
        if not cat_s:
            continue
        out.append(
            (
                cat_s,
                name_s,
                None if desc is None or str(desc).strip() == "" else str(desc).strip(),
                _as_decimal(price_raw),
                None if avail is None or str(avail).strip() == "" else str(avail).strip(),
            )
        )
    return out


def _candidate_menu_items(
    db_session: Session, *, scope_restaurant: bool
) -> list[models.MenuItem]:
    items = (
        db_session.query(models.MenuItem)
        .filter(models.MenuItem.is_active == True)  # noqa: E712
        .order_by(models.MenuItem.id)
        .all()
    )
    out: list[models.MenuItem] = []
    for m in items:
        bar = _is_bar_category(m.category)
        if scope_restaurant and bar:
            continue
        if not scope_restaurant and not bar:
            continue
        out.append(m)
    return out


def _find_match(
    candidates: list[models.MenuItem],
    *,
    category: str,
    name: str,
) -> models.MenuItem | None:
    nn = _norm_name(name)
    nc = _norm_name(category)
    same_both = [
        m for m in candidates if _norm_name(m.name) == nn and _norm_name(m.category) == nc
    ]
    if len(same_both) == 1:
        return same_both[0]
    if len(same_both) > 1:
        return min(same_both, key=lambda m: m.id)

    same_name = [m for m in candidates if _norm_name(m.name) == nn]
    if len(same_name) == 1:
        return same_name[0]
    if len(same_name) > 1:
        same_name_cat = [m for m in same_name if _norm_name(m.category) == nc]
        if len(same_name_cat) == 1:
            return same_name_cat[0]
        return min(same_name, key=lambda m: m.id)
    return None


def _find_match_fallback_all(
    db_session: Session,
    *,
    category: str,
    name: str,
) -> models.MenuItem | None:
    nn = _norm_name(name)
    nc = _norm_name(category)
    items = (
        db_session.query(models.MenuItem)
        .filter(models.MenuItem.is_active == True)  # noqa: E712
        .order_by(models.MenuItem.id)
        .all()
    )
    same_both = [
        m for m in items if _norm_name(m.name) == nn and _norm_name(m.category) == nc
    ]
    if same_both:
        return min(same_both, key=lambda m: m.id)
    same_name = [m for m in items if _norm_name(m.name) == nn]
    if same_name:
        return min(same_name, key=lambda m: m.id)
    return None


def parse_cli(argv: list[str]) -> tuple[Path, bool]:
    """Devuelve (ruta al xlsx, replace_extras)."""
    replace = "--replace" in argv or "-R" in argv
    path_args = [a for a in argv[1:] if not a.startswith("-")]
    if path_args:
        return Path(path_args[0]).expanduser().resolve(), replace
    env = os.getenv("PICNIC_MENU_XLSX")
    if env:
        return Path(env).expanduser().resolve(), replace
    raise SystemExit(
        "Indica la ruta al Excel:\n"
        "  python -m app.sync_menu_picnic_xlsx [--replace] /ruta/menu.xlsx\n"
        "o define PICNIC_MENU_XLSX."
    )


def run_sync(
    xlsx_path: Path,
    *,
    replace_extras: bool = False,
    session: Session | None = None,
) -> dict[str, int]:
    if not xlsx_path.is_file():
        raise FileNotFoundError(xlsx_path)

    own = session is None
    s = session or db.SessionLocal()
    stats = {
        "updated": 0,
        "inserted": 0,
        "unchanged": 0,
        "fallback_match": 0,
        "deactivated": 0,
    }
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        tasks: list[tuple[bool, tuple]] = []
        excel_keys: set[tuple[str, str]] = set()
        for sheet_name, scope_rest in (
            ("Restaurante", True),
            ("Bar", False),
        ):
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"La hoja «{sheet_name}» no existe en el Excel")
            for row in _parse_sheet_rows(wb[sheet_name]):
                tasks.append((scope_rest, row))
                cat_raw, name, _, _, _ = row
                cat_fmt = _format_category(_normalize_category_label(cat_raw))
                excel_keys.add((_norm_name(name), _norm_name(cat_fmt)))

        candidates_rest = _candidate_menu_items(s, scope_restaurant=True)
        candidates_bar = _candidate_menu_items(s, scope_restaurant=False)

        for scope_rest, (cat_raw, name, desc, price, avail) in tasks:
            cand = candidates_rest if scope_rest else candidates_bar
            cat_fmt = _format_category(_normalize_category_label(cat_raw))
            description = _build_description(desc, avail)

            m = _find_match(cand, category=cat_fmt, name=name)
            if m is None:
                m = _find_match_fallback_all(s, category=cat_fmt, name=name)
                if m is not None:
                    stats["fallback_match"] += 1

            if m is None:
                s.add(
                    models.MenuItem(
                        name=name,
                        category=cat_fmt,
                        price=price,
                        description=description,
                        ingredients=None,
                        is_active=True,
                    )
                )
                stats["inserted"] += 1
                continue

            changed = False
            if m.name != name:
                m.name = name
                changed = True
            if m.category != cat_fmt:
                m.category = cat_fmt
                changed = True
            if m.price != price:
                m.price = price
                changed = True
            desc_val = description
            if (m.description or None) != (desc_val or None):
                m.description = desc_val
                changed = True
            if changed:
                stats["updated"] += 1
            else:
                stats["unchanged"] += 1

        if replace_extras:
            for m in (
                s.query(models.MenuItem)
                .filter(models.MenuItem.is_active == True)  # noqa: E712
                .all()
            ):
                if _is_internal_menu_item(m):
                    continue
                key = (_norm_name(m.name), _norm_name(m.category))
                if key in excel_keys:
                    continue
                m.is_active = False
                stats["deactivated"] += 1

        if own:
            s.commit()
        return stats
    except Exception:
        if own:
            s.rollback()
        raise
    finally:
        if own:
            s.close()


def main(argv: list[str] | None = None) -> None:
    argv = argv if argv is not None else sys.argv
    path, replace_extras = parse_cli(argv)
    try:
        out = run_sync(path, replace_extras=replace_extras)
    except OperationalError as err:
        url_hint = os.getenv("DATABASE_URL", "")
        masked = url_hint
        if "@" in masked and "://" in masked:
            head, _, tail = masked.partition("@")
            if "://" in head:
                scheme_user = head.split("://", 1)[0] + "://***:***"
                masked = f"{scheme_user}@{tail}"
        print(
            "\nNo se pudo conectar a PostgreSQL.\n"
            f"DATABASE_URL actual (enmascarada): {masked or '(vacía → se usa el default de db.py, socket local)'}\n\n"
            "Qué hacer:\n"
            "  • Si usas Docker en este repo: levanta la base y usa host y puerto del host, p. ej.\n"
            "    export DATABASE_URL='postgresql://girona_user:girona_pass_change_me@127.0.0.1:25432/girona_prod'\n"
            "    (ajusta usuario, clave y nombre de BD según tu .env en la raíz del proyecto).\n"
            "  • Si la app corre en Railway u otro hosting: pega ahí la misma DATABASE_URL.\n"
            "  • Si Postgres es local instalado: export DATABASE_URL='postgresql://USER:PASS@127.0.0.1:5432/DB'\n",
            file=sys.stderr,
        )
        raise SystemExit(1) from err
    msg = (
        f"Menú sincronizado desde {path} | actualizados: {out['updated']} | "
        f"nuevos: {out['inserted']} | sin cambios: {out['unchanged']} | "
        f"coincidencias fuera de pestaña: {out['fallback_match']}"
    )
    if replace_extras:
        msg += f" | desactivados (no estaban en Excel): {out['deactivated']}"
    print(msg)


if __name__ == "__main__":
    main()
