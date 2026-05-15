"""
Carga deduplicada: ESTANDARIZACION (menú+recetas), LISTA PROVEEDORES (inventario costos),
Inventario cocina (materiales), menu_escrito (insumos extra + bar + cristalería),
costo salsas (recetas internas inactivas).

(Las fotos en IMAGENES_MENU sirven solo como referencia humana al interpretar el menú
escrito; no se almacenan ni se muestran en la app.)

ATENCIÓN: al completarse con éxito, este script borra y vuelve a cargar inventario,
menú, ventas POS, etc. No ejecutarlo en producción si necesitas conservar ítems o
ventas creados en la app (o haz copia de la base antes).

Uso (desde la carpeta girona-back, con .env o DATABASE_URL):

  export GIRONA_DATA_DIR="/home/hp/Documents/GIRONA"
  python -m app.seed_girona_data

O con Docker:

  docker compose run --rm -e GIRONA_DATA_DIR=/data -v /home/hp/Documents/GIRONA:/data:ro \
    backend python -m app.seed_girona_data
"""
from __future__ import annotations

import os
import re
import unicodedata
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from . import db, models
from .recetario_bar_seed import apply_recetario_bar_items, sync_recetario_bar_recipes

# Orden: tablas hoja adentro primero
_DELETE_SQL = text(
    """
    DELETE FROM electronic_invoices;
    DELETE FROM sale_items;
    DELETE FROM sales;
    DELETE FROM pos_order_items;
    DELETE FROM pos_orders;
    DELETE FROM pos_tables;
    DELETE FROM recipe_items;
    DELETE FROM recipes;
    DELETE FROM stock_movements;
    DELETE FROM purchase_items;
    DELETE FROM purchases;
    DELETE FROM menu_items;
    DELETE FROM inventory_products;
    DELETE FROM suppliers;
    """
)


def _norm(s: str) -> str:
    t = " ".join((s or "").upper().split())
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t


def _find_girona_dir() -> Path:
    env = os.getenv("GIRONA_DATA_DIR")
    if env:
        return Path(env).resolve()
    here = Path(__file__).resolve().parent.parent
    cands = [
        here / "girona_seed_data",
        Path.home() / "Documents" / "GIRONA",
    ]
    for c in cands:
        if c.is_dir() and (c / "ESTANDARIZACION JOSE GIRONA.xlsx").exists():
            return c
    return cands[-1]


def _f(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(" ", "")
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _assign_category(lineno: int, name: str) -> str:
    n = name.upper()
    if any(k in n for k in ("CREPE", "WAFLE", "CUAJADA EN REDUCCION", "DULCE TENTACION", "DULCE ")):
        return "Postres"
    if lineno < 126:
        if "ENSALADA" in n or "VINAGRETA" in n:
            return "Ensaladas"
        return "Entradas"
    if 126 <= lineno < 721:
        return "Platos fuertes"
    if 721 <= lineno < 802:
        return "Platos especiales"
    if lineno >= 802 and lineno < 838:
        return "Menú infantil"
    return "Postres"


def _parse_estandarizacion_rows(rows: list[tuple]) -> list[dict[str, Any]]:
    n = len(rows)
    starts: list[int] = []
    for i, row in enumerate(rows):
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        if not a or b is None:
            continue
        sa = str(a).strip()
        if ("NOMBRE DEL PRODUCTO" in sa or "NOMBRE DEL PLATO" in sa) and str(b).strip():
            starts.append(i)
    out: list[dict[str, Any]] = []
    for bi, s in enumerate(starts):
        e = starts[bi + 1] if bi + 1 < len(starts) else n
        block = rows[s:e]
        name = str(block[0][1]).strip()
        ings: list[dict] = []
        block_total: float | None = None
        for r in block[1:]:
            a = r[0] if len(r) > 0 else None
            a_str = str(a).strip() if a is not None else ""
            u = a_str.upper()
            if a_str and ("NOMBRE DEL" in u):
                break
            if a_str in ("PLATOS FUERTES", "PLATOS ESPECIALES", "MENU INFANTIL"):
                continue
            b = r[1] if len(r) > 1 else None
            c = r[2] if len(r) > 2 else None
            d = r[3] if len(r) > 3 else None
            col_e = r[4] if len(r) > 4 else None
            if b and "TOTAL" in str(b).upper() and col_e is not None:
                try:
                    block_total = float(col_e)
                except (TypeError, ValueError):
                    pass
            if d and "TOTAL" in str(d).upper() and col_e is not None:
                try:
                    block_total = float(col_e)
                except (TypeError, ValueError):
                    pass
            if not a_str:
                if col_e is not None and isinstance(col_e, (int, float)) and c is None and d is None and b is None and a is None:
                    if 0 < float(col_e) < 1_000_000_000:
                        block_total = float(col_e)
                continue
            if u.startswith("INGREDIENTE"):
                continue
            if "GRAMAJE" in u or "PRODUCCION" in u or u.startswith("PRECIO") and "GR" in u:
                continue
            if "NOMBRE" in u:
                continue
            if a_str in ("RINDE", "TOTAL", "SAL", "PIMIENTA") and c is None and not isinstance(d, (int, float)):
                if not c and not d:
                    continue
            iname = a_str
            if not iname or len(iname) < 1:
                continue
            ings.append(
                {
                    "name": iname,
                    "unit": (str(b).strip() if b is not None else "") or "ud",
                    "weight": _f(c),
                    "price": _f(d),
                    "total": _f(col_e) if col_e is not None else None,
                }
            )
        if block_total and block_total > 1_000_000:
            block_total = None
        out.append(
            {
                "name": name,
                "line": s + 1,
                "ingredients": ings,
                "food_cost": block_total,
            }
        )
    return out


def _price_from_cost(food_cost: float | None) -> Decimal:
    if food_cost is None or food_cost <= 0:
        return Decimal("15000")
    if food_cost > 1_000_000:
        return Decimal("15000")
    p = int(food_cost * 2.2)
    p = max(8_000, min(400_000, p))
    return Decimal(str(p))


def _parse_salsas(wb) -> list[dict[str, Any]]:
    ws = wb["costo salsas girona "]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    out: list[dict[str, Any]] = []
    i = 0
    while i < len(rows):
        a = rows[i][0] if len(rows[i]) else None
        if not a or not str(a).strip():
            i += 1
            continue
        a0 = str(a).strip()
        u0 = a0.upper()
        if u0 in ("INGREDIENTES",) or "UNIDAD" in u0 and "MEDIDA" in u0:
            i += 1
            continue
        if i + 1 < len(rows) and str(rows[i + 1][0] or "").strip().upper() == "RINDE":
            name = a0
            rinde = rows[i + 1][1] if len(rows[i + 1]) > 1 else None
            i += 2
            if i < len(rows) and str(rows[i][0] or "").upper().startswith("INGREDIENT"):
                i += 1
            ings: list[dict] = []
            while i < len(rows):
                r = rows[i]
                ra = r[0] if len(r) else None
                if not ra or not str(ra).strip():
                    t = r[2] if len(r) > 2 else None
                    if t and str(t).lower() == "total":
                        break
                    i += 1
                    continue
                sra = str(ra).upper()
                if sra == "RINDE":
                    i -= 1
                    break
                if sra == "INGREDIENTES" or sra.startswith("UNIDAD"):
                    i += 1
                    continue
                b = r[1] if len(r) > 1 else None
                c = r[2] if len(r) > 2 else None
                d = r[3] if len(r) > 2 else None
                if c is None and d is None:
                    i += 1
                    continue
                ings.append(
                    {
                        "name": str(ra).strip(),
                        "unit": str(b or "").strip() or "ud",
                        "weight": _f(c),
                        "price": _f(d) if d is not None else Decimal("0"),
                        "total": _f(d) if d is not None else None,
                    }
                )
                i += 1
            out.append({"name": f"[Insumo] {name}", "line": 0, "ingredients": ings, "food_cost": None, "rinde": rinde})
            continue
        i += 1
    return out


def _load_lista_proveedores(wb) -> list[tuple[str, Decimal, str | None]]:
    ws = wb["LISTA DE PROVEEDORES"]
    by_norm: dict[str, tuple[str, Decimal, str | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        prod = row[1]
        if not prod or not str(prod).strip():
            continue
        name = " ".join(str(prod).split())
        pack = _f(row[2])
        val = _f(row[4])  # $/g o $/ml
        u = f"ref {pack} u" if pack else "ud"
        key = _norm(name)
        if key in by_norm:
            continue
        by_norm[key] = (name, val, u)
    return list(by_norm.values())


def _load_inventario_cocina(wb) -> list[tuple[str, str, float]]:
    ws = wb["Inventario Cocina"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat, desc, qty = row[0], row[1], row[2]
        if not desc or not str(desc).strip():
            continue
        out.append(
            (str(cat or "Cocina").strip(), " ".join(str(desc).split()), float(qty or 0))
        )
    return out


def _parse_menu_text(
    path: Path,
) -> tuple[set[str], list[tuple[str, str, float]], list[tuple[str, str]]]:
    """
    Cocina: líneas 2-93. Bar: 94-132. Equipo 133-155. Cristalería: bloque '21 de enero 2026'.
    """
    raw = path.read_text(encoding="utf-8", errors="replace").splitlines()
    ingredients: set[str] = set()
    cristal: list[tuple[str, str, float]] = []
    equip: list[tuple[str, str, float]] = []
    bar: list[tuple[str, str]] = []
    mode: str = "body"
    for i, line in enumerate(raw, start=1):
        t = " ".join(line.split()).strip()
        if not t or t == "MENÚ":
            continue
        if "Inventario cristalería" in t:
            mode = "cristal" if "21 de enero" in t else "cristal_drop"
            continue
        if mode == "cristal_drop":
            continue
        if mode == "cristal":
            m = re.match(r"^(.+?)\s+(\d+)\s*$", t)
            if m:
                cristal.append((m.group(1).strip(), "Cristalería", float(m.group(2))))
            continue
        if 2 <= i <= 93:
            ingredients.add(t)
        elif 94 <= i <= 132:
            bar.append(_bar_categorize(t))
        elif 133 <= i <= 155:
            equip.append((t, "Almacén", 1.0))
    if cristal:
        seen: dict[str, float] = {}
        for n, c, q in cristal:
            seen[n] = max(seen.get(n, 0), q)
        cristal = [(k, "Cristalería", v) for k, v in seen.items()]
    return ingredients, cristal + equip, bar


def _bar_categorize(name: str) -> tuple[str, str]:
    t = " ".join(name.split())
    n = t.lower()
    if any(k in n for k in ("cerveza", "poker", "heineken", "stella", "aguila", "coron", "club col", "3 cordill", "postob", "ginet", "andina", "segu ole")):
        return ("Cervezas nacionales", t)
    if any(k in n for k in ("ginebra", "tequila", "ron", "vino", "aguard", "champag", "triple", "soda tón", "coca", "gaseos", "blue", "tónica", "cuatro")):
        if "vino" in n:
            return ("Vinos", t)
        if any(k in n for k in ("coca", "gaseos", "ton", "soda", "hola", "andina", "segu", "ginet", "postob", "manzana", "uva", "cola y pola")):
            return ("Sodas", t)
        return ("Cocteleria", t)
    if any(k in n for k in ("helado", "gomit", "banderit", "chispa", "leche conden", "tajín", "chantilly")):
        return ("Postres", t)
    return ("Sodas", t) if t else ("Sodas", t)


def _match_product(ing_name: str, by_key: dict[str, int]) -> int | None:
    k = _norm(ing_name)
    if k in by_key:
        return by_key[k]
    for k2, pid in by_key.items():
        if k in k2 or k2 in k:
            return pid
    return None


def run_seed(
    data_dir: Path | None = None, session: Session | None = None
) -> None:
    base = (data_dir or _find_girona_dir()).resolve()
    ex_path = base / "ESTANDARIZACION JOSE GIRONA.xlsx"
    inv_path = base / "Inventario_Cocina_Chef_Jose_Villarreal.xlsx"
    txt_path = base / "menu_escrito.txt"
    if not ex_path.exists():
        raise FileNotFoundError(f"Falta {ex_path}")
    if not inv_path.exists():
        raise FileNotFoundError(f"Falte {inv_path}")
    if not txt_path.exists():
        raise FileNotFoundError(f"Falte {txt_path}")

    own_session = session is None
    db_session = session or db.SessionLocal()
    try:
        # Un solo commit al final: si algo falla antes, rollback() revierte también
        # el DELETE y no se pierde el menú ni datos creados desde la app.
        db_session.execute(_DELETE_SQL)

        wb = openpyxl.load_workbook(ex_path, data_only=True)
        ws = wb["Copia de ESTANDARIZACION "]
        all_rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        blocks = _parse_estandarizacion_rows(all_rows)
        lista = _load_lista_proveedores(wb)
        salsas = _parse_salsas(wb)
        inv_c = _load_inventario_cocina(
            openpyxl.load_workbook(inv_path, data_only=True)
        )
        xtra_ing, materials, bar_items = _parse_menu_text(txt_path)

        inv_by_name: dict[str, models.InventoryProduct] = {}
        sku = 0

        by_key: dict[str, int] = {}

        def _add_inv(name: str, unit: str | None, cost: Decimal, kind: str) -> models.InventoryProduct:
            nonlocal sku
            k = _norm(name)
            if k in inv_by_name:
                return inv_by_name[k]
            sku += 1
            p = models.InventoryProduct(
                name=name,
                sku=f"INV-{sku:05d}",
                kind=kind,
                unit=unit,
                on_hand=Decimal("0"),
                average_cost=cost,
                last_cost=cost,
                is_active=True,
            )
            db_session.add(p)
            db_session.flush()
            inv_by_name[k] = p
            by_key[_norm(p.name)] = p.id
            return p

        for name, val, udesc in lista:
            _add_inv(name, udesc, val if val else Decimal("0"), "ingredient")

        for s in xtra_ing:
            if _norm(s) in inv_by_name:
                continue
            _add_inv(s, "ud", Decimal("0"), "ingredient")

        for _cat, desc, row_qty in inv_c + materials:
            k = _norm(desc)
            if k in inv_by_name:
                o = inv_by_name[k]
                o.on_hand = Decimal(str(row_qty))
                o.kind = "material"
                continue
            sku += 1
            p = models.InventoryProduct(
                name=desc,
                sku=f"MAT-{sku:05d}",
                kind="material",
                unit="ud",
                on_hand=Decimal(str(row_qty)),
                average_cost=Decimal("0"),
                last_cost=Decimal("0"),
                is_active=True,
            )
            db_session.add(p)
            db_session.flush()
            inv_by_name[k] = p
            by_key[k] = p.id

        for p in inv_by_name.values():
            by_key[_norm(p.name)] = p.id

        for b in blocks:
            cat = _assign_category(b["line"], b["name"])
            desc_bits = []
            if b.get("food_cost"):
                desc_bits.append(
                    f"Costo estimado hoja: ${b['food_cost']:.2f} COP (referencia)."
                )
            desc = " ".join(desc_bits) or None
            ings = b["ingredients"]
            norm_ings: list[dict] = []
            for ing in ings:
                w = _f(ing.get("weight")) if ing.get("weight") is not None else Decimal("0")
                if w <= 0:
                    w = Decimal("1")
                norm_ings.append(
                    {
                        "name": ing["name"],
                        "unit": (ing.get("unit") or "ud")[:20],
                        "weight": float(w),
                        "price": float(_f(ing.get("price"))),
                        "total": float(_f(ing["total"]))
                        if ing.get("total") is not None
                        else None,
                    }
                )
            price = _price_from_cost(b.get("food_cost"))
            m = models.MenuItem(
                name=" ".join(b["name"].split()),
                category=cat,
                price=price,
                description=desc,
                ingredients=norm_ings,
                is_active=True,
            )
            db_session.add(m)
            db_session.flush()
            r = models.Recipe(
                menu_item_id=m.id,
                yield_quantity=Decimal("1"),
                unit="porcion",
            )
            db_session.add(r)
            db_session.flush()
            for ing in ings:
                pid = _match_product(ing["name"], by_key)
                if not pid:
                    np = _add_inv(ing["name"], ing.get("unit") or "ud", Decimal("0"), "ingredient")
                    pid = np.id
                wv = _f(ing.get("weight")) if ing.get("weight") is not None else Decimal("0")
                r_qty = wv if wv > 0 else Decimal("1")
                db_session.add(
                    models.RecipeItem(
                        recipe_id=r.id,
                        product_id=pid,
                        quantity=r_qty,
                        waste_pct=Decimal("0"),
                    )
                )

        for b in salsas:
            name = b["name"]
            ings = b["ingredients"]
            s_norm: list[dict] = []
            for ing in ings:
                w = _f(ing.get("weight")) if ing.get("weight") is not None else Decimal("0")
                if w <= 0:
                    w = Decimal("1")
                s_norm.append(
                    {
                        "name": ing["name"],
                        "unit": (ing.get("unit") or "ud")[:20],
                        "weight": float(w),
                        "price": float(_f(ing.get("price"))),
                        "total": float(_f(ing["total"]))
                        if ing.get("total") is not None
                        else None,
                    }
                )
            m = models.MenuItem(
                name=name,
                category="Insumos",
                price=Decimal("0"),
                description="Receta interna (hoja salsas).",
                ingredients=s_norm,
                is_active=False,
            )
            db_session.add(m)
            db_session.flush()
            r = models.Recipe(menu_item_id=m.id, yield_quantity=Decimal("1"), unit="lote", notes="costo salsas")
            db_session.add(r)
            db_session.flush()
            for ing in ings:
                pid = _match_product(ing["name"], by_key)
                if not pid:
                    np = _add_inv(ing["name"], ing.get("unit") or "ud", Decimal("0"), "ingredient")
                    pid = np.id
                wv = _f(ing.get("weight")) if ing.get("weight") is not None else Decimal("0")
                r_qty = wv if wv > 0 else Decimal("1")
                db_session.add(
                    models.RecipeItem(
                        recipe_id=r.id,
                        product_id=pid,
                        quantity=r_qty,
                        waste_pct=Decimal("0"),
                    )
                )

        for cat, tname in bar_items:
            if not tname:
                continue
            k2 = _norm(tname)
            dup = (
                db_session.query(models.MenuItem)
                .filter(func.lower(models.MenuItem.name) == tname.lower().strip())
                .first()
            )
            if dup:
                continue
            m = models.MenuItem(
                name=tname,
                category=cat,
                price=Decimal("8000"),
                description="Incluido en lista escrita (revisar precio).",
                ingredients=None,
                is_active=True,
            )
            db_session.add(m)

        apply_recetario_bar_items(db_session)
        # Flush para que sync vea Recipes creados en apply y no duplique ix_recipes_menu_item_id.
        db_session.flush()
        sync_recetario_bar_recipes(db_session)

        db_session.commit()
        if own_session:
            db_session.close()
    except Exception:
        if own_session:
            db_session.rollback()
            db_session.close()
        raise


def main() -> None:
    run_seed()


if __name__ == "__main__":
    main()
